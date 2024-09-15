import openpyxl
from openpyxl.styles import PatternFill

def get_rows_from_sheet(file_path, sheet_name, *column_indices):
    """
    Reads data from the specified columns of an Excel sheet, ignoring the header row.
    
    Args:
        file_path (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to read from.
        column_indices (int): The indices of the columns to read (1-based).

    Returns:
        list: A list of tuples with values from the specified columns, excluding the header.
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    rows = []
    
    # Skip the header (first row) and get the column data
    for row in sheet.iter_rows(min_row=2, values_only=True):
        selected_values = tuple(row[col_index - 1] for col_index in column_indices)
        rows.append(selected_values)
    return rows, wb

def check_row_in_other_sheet(row_values, rows_b):
    """
    Check if a tuple of values exists as a row in another list of rows.
    
    Args:
        row_values (tuple): The tuple of values to check.
        rows_b (list): The list of rows to check against.

    Returns:
        bool: True if the row exists in rows_b, False otherwise.
    """
    return row_values in rows_b

def highlight_entire_row(sheet, row_idx, max_col, color):
    """
    Highlights an entire row by changing the background color of all cells with values.
    
    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to modify.
        row_idx (int): The row index that should be highlighted.
        max_col (int): The maximum number of columns to highlight.
        color (str): The color code for highlighting (e.g., '00FF00' for green, 'FF0000' for red).
    """
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    for col_idx in range(1, max_col + 1):
        cell = sheet.cell(row=row_idx + 2, column=col_idx)  # +2 because row_idx is 0-based, and we skip header
        if cell.value:  # Only highlight cells with values
            cell.fill = fill

def main():
    # File paths and sheet names
    file_a = 'FISUB.xlsx'
    file_b = 'ENT2_CallLog.xlsx'
    sheet_name_a = 'Sheet1'  # Update as necessary
    sheet_name_b = 'Sheet1'  # Update as necessary
    cTypeA = 2  # Column to compare from Spreadsheet A (1-based index)
    cTypeB = 2  # Column to compare from Spreadsheet B (1-based index)
    SubA = 1  # Another column to compare from Spreadsheet A (1-based index)
    SubB = 1  # Another column to compare from Spreadsheet B (1-based index)

    # Read specific columns data from both spreadsheets, excluding headers
    rows_a, wb_a = get_rows_from_sheet(file_a, sheet_name_a, SubA, cTypeA)
    rows_b, _ = get_rows_from_sheet(file_b, sheet_name_b, SubB, cTypeB)
    sheet_a = wb_a[sheet_name_a]

    # Get the maximum number of columns to highlight for each row
    max_col = sheet_a.max_column

    # Track matching and missing rows
    matching_rows = []
    missing_rows = []
    
    for i, row in enumerate(rows_a):
        if check_row_in_other_sheet(row, rows_b):
            matching_rows.append(i)
            print(f"Row {row} exists in both Spreadsheet A and B.")
        else:
            missing_rows.append(i)
            print(f"Row {row} is not in Spreadsheet B.")

    # Highlight entire matching rows in green and missing rows in red in Spreadsheet A
    for row_idx in matching_rows:
        highlight_entire_row(sheet_a, row_idx, max_col, "00FF00")  # Green for matches

    for row_idx in missing_rows:
        highlight_entire_row(sheet_a, row_idx, max_col, "FF0000")  # Red for missing rows

    # Save the modified Spreadsheet A
    wb_a.save('FISUB_highlighted.xlsx')

if __name__ == "__main__":
    main()


