import openpyxl
import os

# Directory containing Excel files
directory = 'C:\\Enterprise Mechanical\\FISUB\\FISUB_2023'  # Replace with the path to your directory

# Create a new workbook to hold combined data
combined_wb = openpyxl.Workbook()
combined_sheet = combined_wb.active
combined_sheet.title = "Combined Data"

# Initialize row counter and sheet counter
current_row = 1
sheet_counter = 1
first_file = True

# Function to create a new sheet if the row limit is exceeded
def create_new_sheet(wb, counter):
    new_sheet = wb.create_sheet(title=f"Combined Data {counter}")
    return new_sheet

# Iterate through all files in the directory
for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(directory, filename)
        
        # Load the workbook and process each sheet
        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception as e:
            print(f"Failed to load workbook {filename}: {e}")
            continue
        
        for sheet_name in wb.sheetnames:
            print(f"Processing sheet: {sheet_name} from file: {filename}")
            sheet = wb[sheet_name]
            
            # Handle headers, only for the first sheet
            if first_file:
                # Add a header for the filename column
                combined_sheet.cell(row=current_row, column=1).value = "Filename"
                header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                for col_num, header in enumerate(header_row, start=2):
                    combined_sheet.cell(row=current_row, column=col_num).value = header
                current_row += 1
                first_file = False
            
            # Copy data with filename in the first column
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(cell is not None for cell in row):
                    continue  # Skip empty rows

                if current_row > 1048576:
                    # Create a new sheet if row limit is reached
                    sheet_counter += 1
                    combined_sheet = create_new_sheet(combined_wb, sheet_counter)
                    current_row = 1
                    # Add the filename header to the new sheet
                    combined_sheet.cell(row=current_row, column=1).value = "Filename"
                    header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                    for col_num, header in enumerate(header_row, start=2):
                        combined_sheet.cell(row=current_row, column=col_num).value = header
                    current_row += 1

                try:
                    combined_sheet.cell(row=current_row, column=1).value = filename
                    for col_num, value in enumerate(row, start=2):
                        combined_sheet.cell(row=current_row, column=col_num).value = value
                    current_row += 1
                except Exception as e:
                    print(f"Error writing row {current_row} from sheet {sheet_name} in file {filename}: {e}")

# Save the combined workbook
combined_wb.save('combined_data_V8.xlsx')

# Notify that the process is complete
print("All sheets have been processed and combined into 'combined_data_V8.xlsx'.")

